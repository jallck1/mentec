import os
from openpyxl import load_workbook
from rest_framework import viewsets
from . import models, serializers
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.models import Group


class ProductsView(viewsets.ModelViewSet):
    queryset = models.Productos.objects.all()
    serializer_class = serializers.ProductSerializer
    
    def get_all_products(self, request):
        products = models.Productos.objects.all()
        serializer = serializers.ProductSerializer(products, many=True)
        return JsonResponse({'data':serializer.data})  
    
    def create_product(self, request):
        serializer = serializers.ProductSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return JsonResponse({}, status=201) 
    
    def update_product(self, request, id):
        product = get_object_or_404(models.Productos, id=id)
        serializer = serializers.ProductSerializer(product,data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return JsonResponse({}, status=204)
    
    def delete_product(self, request, id):
        product = get_object_or_404(models.Productos, id=id)
        product.delete()
        return JsonResponse({}, status=204)
    
class UsersView(viewsets.ModelViewSet):
    queryset = models.User.objects.all()
    serializer_class = serializers.UserSerializer
    
    def get_users(self, request):
        queryset = models.User.objects.all()
        serializer = serializers.UserSerializer(queryset, many=True)
        return JsonResponse({'data':serializer.data}, status=200)
    
    
    def mostrar_mensaje(self, request, name):
        return JsonResponse({'mensaje':f'Hola {name}'})

    def get_user_balance(self, request, id):
        user = get_object_or_404(models.User, id=id)
        return JsonResponse({'balance':user.balance})
    
    def create_user(self, request):
        group = Group.objects.get(name='Compradores')
        name = request.data.get('name',None)
        password = request.data.get('password',None)
        if not name or not password:
            return JsonResponse({'msg':'Deposite bien los datos'}, status=400)
        else:
            user = models.User(name=name)
            user.set_password(password)
            user.save()
            user.groups.add(group)
            user.save()
        return JsonResponse({},status=204)
    
    def desactivate_user(self, request,id):
        user = get_object_or_404(models.User, id=id)
        user.is_active = False
        user.save()
        return JsonResponse({}, status=204)
    


class ClientView(viewsets.ModelViewSet):
    queryset = models.Cliente.objects.all()
    serializer_class = serializers.ClientViewSerializer
    
    def get_all_clients(self, request):
        clientes = models.Cliente.objects.all()
        serializer = serializers.ClientViewSerializer(clientes, many=True)
        return JsonResponse({'data':serializer.data})
    
    def post_clients(self, request):
        print(request.data)
        serializer = serializers.ClientViewSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return JsonResponse({'msg':'Creado con exito'}, status=201) 
    
    def delete_cliente(self, request, id):
        
        product = get_object_or_404(models.Cliente, id=id)
        product.delete()
        return JsonResponse({}, status=204)
        
        
class TransactsView(viewsets.ModelViewSet):
    queryset = models.Transacts.objects.all()
    serializer_class = serializers.TransactSerializer

    def get_all_transacts(self, request):
        transacts= models.Transacts.objects.all()
        serializer = serializers.TransactSerializer(transacts, many=True)
        return JsonResponse({'data':serializer.data})

    def buy_product(self, request, product_id, buyer_id):
        product = get_object_or_404(models.Productos, id=product_id)
        user = get_object_or_404(models.User, id=buyer_id)
        transact = models.Transacts.objects.create(product=product, buyer=user, quantity=1, total=product.price)
        user.balance -= transact.total
        user.save()
        return JsonResponse({}, status=201)
    
    def get_transacts_user_based(self, request,id_user):
        user= get_object_or_404(models.User,id=id_user)
        transacts = models.Transacts.objects.filter(buyer=user)
        serializer = serializers.TransactSerializer(transacts, many=True)
        return JsonResponse({'data':serializer.data})
    
    def get_transacts_xlsx(self, request):
        transacts = models.Transacts.objects.all()
        serializer = serializers.TransactSerializer(transacts, many=True)
        data = serializer.data
        print(os.path.abspath('dat_pl.xlsx'))
        workbook = load_workbook(os.path.abspath('dat_pl.xlsx'))
        worksheet = workbook['Hoja1']
        for row, item in enumerate(data, start=2):
            worksheet.cell(row=row, column=1, value=item['id'])
            worksheet.cell(row=row, column=2, value=item['product'])
            worksheet.cell(row=row, column=3, value=item['buyer'])
            worksheet.cell(row=row, column=4, value=item['createdAt'])
            worksheet.cell(row=row, column=5, value=item['quantity'])
            worksheet.cell(row=row, column=6, value=float(item['total']))
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="archivo.xlsx"'
        workbook.save(response)
        return response
    
class ImageView(viewsets.ModelViewSet):

    def get_file_img(self, request, image_name):
        pathFile = os.path.abspath(__file__)
        path = pathFile.replace(r"MAIN\views.py", "") + r"images" + '\\' + image_name   
        print(path)
        if os.path.exists(path=path):
            return FileResponse(open(path,'rb'),content_type='image/jpeg')
        else:
            print('ruta')
            return JsonResponse({'error':'No se encuentra la imagen'}, status=404)